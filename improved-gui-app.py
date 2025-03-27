import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import subprocess
import platform
from docx_to_excel_processor import DocxToExcelProcessor
import openpyxl

# Импортируем класс для обработки адресов
from b_column_parser import ImprovedAddressProcessor

class DocxToExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Таблицы из DOCX в Excel")
        self.root.geometry("600x500")
        
        # Переменные для хранения путей к файлам
        self.docx_path = None
        self.excel_path = None
        
        # Создаем экземпляр обработчика
        self.processor = DocxToExcelProcessor()
        self.address_processor = ImprovedAddressProcessor()
        
        # Создание интерфейса
        self.create_gui()
    
    def create_gui(self):
        """Создание графического интерфейса"""
        # Главный фрейм
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Заголовок
        title_label = ttk.Label(main_frame, text="Конвертация таблиц из DOCX в Excel", font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # Фрейм для выбора файла
        file_frame = ttk.LabelFrame(main_frame, text="Выбор DOCX файла", padding=10)
        file_frame.pack(fill=tk.X, pady=10)
        
        # Поле для отображения пути к файлу
        self.file_path_var = tk.StringVar()
        file_path_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=50)
        file_path_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        # Кнопка выбора файла
        browse_button = ttk.Button(file_frame, text="Обзор...", command=self.select_file)
        browse_button.grid(row=0, column=1, padx=5, pady=5)
        
        file_frame.columnconfigure(0, weight=1)
        
        # Информационное поле
        info_frame = ttk.LabelFrame(main_frame, text="Статус", padding=10)
        info_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Текстовое поле для статуса
        self.status_text = tk.Text(info_frame, wrap=tk.WORD, height=8)
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Полоса прокрутки
        scrollbar = ttk.Scrollbar(info_frame, orient="vertical", command=self.status_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.configure(yscrollcommand=scrollbar.set)
        
        self.update_status("Выберите DOCX файл для начала обработки.")
        
        # Фрейм для кнопок
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        # Кнопка обработки
        process_button = ttk.Button(
            button_frame, 
            text="Обработать", 
            command=self.process_file
        )
        process_button.pack(side=tk.LEFT, padx=5)
        
        # Кнопка показа столбца I
        show_column_button = ttk.Button(
            button_frame,
            text="Показать столбец I",
            command=self.show_column_i
        )
        show_column_button.pack(side=tk.LEFT, padx=5)
        
        # Кнопка выхода
        exit_button = ttk.Button(button_frame, text="Выход", command=self.root.destroy)
        exit_button.pack(side=tk.RIGHT, padx=5)
    
    def show_column_i(self):
        """Показывает первые 10 строк столбца I из Excel файла"""
        if not self.excel_path:
            messagebox.showerror("Ошибка", "Сначала выберите и обработайте DOCX файл")
            return
        
        try:
            # Открываем Excel файл
            workbook = openpyxl.load_workbook(self.excel_path)
            
            # Получаем первый лист
            sheet = workbook.active
            
            # Формируем текст для отображения
            result_text = "Первые 10 строк столбца I:\n\n"
            
            # Получаем первые 10 строк из столбца I (индекс 9)
            for row in range(1, min(11, sheet.max_row + 1)):
                cell_value = sheet.cell(row=row, column=9).value
                result_text += f"Строка {row}: {cell_value}\n"
            
            # Показываем результат в отдельном окне
            preview_window = tk.Toplevel(self.root)
            preview_window.title("Предпросмотр столбца I")
            preview_window.geometry("1200x400")
            
            # Создаем текстовое поле с прокруткой
            text_frame = ttk.Frame(preview_window)
            text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            text_widget = tk.Text(text_frame, wrap=tk.WORD)
            text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            text_widget.configure(yscrollcommand=scrollbar.set)
            
            # Вставляем текст
            text_widget.insert(tk.END, result_text)
            text_widget.configure(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть Excel файл: {str(e)}")
    
    def update_status(self, message):
        """Обновление статуса в текстовом поле"""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.insert(tk.END, message)
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()
    
    def select_file(self):
        """Выбор DOCX файла"""
        file_path = filedialog.askopenfilename(
            title="Выберите DOCX файл",
            filetypes=[("DOCX файлы", "*.docx"), ("Все файлы", "*.*")]
        )
        if file_path:
            self.docx_path = file_path
            self.file_path_var.set(file_path)
            
            # Создаем путь для сохранения Excel
            base_name = os.path.basename(file_path)
            name, _ = os.path.splitext(base_name)
            self.excel_path = os.path.join(os.path.dirname(file_path), f"{name}.xlsx")
            
            self.update_status(f"Выбран файл: {file_path}\n\nНажмите кнопку 'Обработать' для конвертации таблиц и обработки данных.")
    
    def process_file(self):
        """Обработка файла: конвертация DOCX в Excel и применение всех правил обработки"""
        if not self.docx_path:
            messagebox.showerror("Ошибка", "Сначала выберите DOCX файл")
            return
        
        try:
            self.update_status("Обработка файла...\nШаг 1: Извлечение таблиц из DOCX...\n\n"
                            "Правила обработки:\n"
                            "- Все таблицы из Word перенесутся в Excel\n"
                            "- Первая строка удаляется, если во второй ячейке НЕТ даты\n"
                            "- Первая строка сохраняется, если во второй ячейке ЕСТЬ дата\n"
                            "- Столбцы A и C будут удалены\n"
                            "- Даты во втором столбце будут приведены к формату ДД.ММ.ГГГГ\n"
                            "- Даты рождения в пятом столбце будут приведены к формату ДД.ММ.ГГГГ\n"
                            "- Даты окончания в восьмом столбце будут приведены к формату ДД.ММ.ГГГГ\n"
                            "- Информация о судах из столбцов D и E будет перемещена в столбец I\n"
                            "- Даты в информации о судах будут нормализованы, включая даты с пробелами")
            
            # Шаг 1: Извлечение таблиц из DOCX в Excel
            table_count = self.processor.convert_docx_to_excel(self.docx_path, self.excel_path)
            
            if table_count > 0:
                self.update_status(f"Шаг 1: Таблицы успешно извлечены ({table_count} шт.)\n\n"
                                "Шаг 2: Удаление столбцов A и C, обработка дат и информации о судах...")
                
                # Шаг 2: Обработка Excel-файла (все операции в одной функции)
                stats = self.processor.process_excel_file(self.excel_path)
                
                # Шаг 3: Дополнительная обработка столбца B - извлечение адресов, телефонов и другой информации
                self.update_status(f"Шаг 3: Обработка столбца B - извлечение адресов, телефонов и другой информации...")
                
                # Загружаем файл для обработки столбца B
                workbook = openpyxl.load_workbook(self.excel_path)
                sheet = workbook.active
                
                # Статистика для столбца B
                b_stats = {
                    'processed_rows': 0,
                    'addresses_found': 0,
                    'phones_found': 0,
                    'other_info_found': 0
                }
                
                # Обрабатываем каждую строку в столбце B
                for row in range(1, sheet.max_row + 1):
                    cell_b = sheet.cell(row=row, column=2)  # Столбец B
                    
                    # Если ячейка не пустая, обрабатываем ее
                    if cell_b.value and str(cell_b.value).strip():
                        raw_text = str(cell_b.value).strip()
                        b_stats['processed_rows'] += 1
                        
                        # 1. Сначала извлекаем телефон
                        phone, original_phone_texts = self.address_processor.extract_phone(raw_text)
                        
                        # 2. Удаляем телефон из исходного текста
                        text_without_phone = raw_text
                        if original_phone_texts:
                            for phone_text in original_phone_texts:
                                text_without_phone = text_without_phone.replace(phone_text, '')
                            b_stats['phones_found'] += 1
                        
                        # 3. Извлекаем адрес из текста без телефона
                        formatted_address, original_address = self.address_processor.extract_address(text_without_phone)
                        
                        # 4. Определяем иное - всё, что осталось после удаления телефона и адреса
                        other_info = text_without_phone
                        if original_address:
                            other_info = other_info.replace(original_address, '')
                            b_stats['addresses_found'] += 1
                        
                        # Очищаем результат
                        other_info = self.address_processor.clean_other_info(other_info)
                        if other_info:
                            b_stats['other_info_found'] += 1
                        
                        # Получаем ячейки в столбцах O, P, Q
                        cell_o = sheet.cell(row=row, column=15)  # Столбец O для адреса
                        cell_p = sheet.cell(row=row, column=16)  # Столбец P для телефона
                        cell_q = sheet.cell(row=row, column=17)  # Столбец Q для иного
                        
                        # Заполняем ячейки
                        if formatted_address:
                            cell_o.value = formatted_address
                        
                        if phone:
                            cell_p.value = phone
                        
                        if other_info:
                            cell_q.value = other_info
                
                # Сохраняем изменения
                workbook.save(self.excel_path)
                
                # Финальное сообщение
                self.update_status(
                    f"Обработка успешно завершена!\n\n"
                    f"- Извлечено таблиц: {table_count}\n"
                    f"- Обработано листов: {stats['sheets_processed']}\n"
                    f"- Удалено первых строк: {stats['rows_deleted']}\n"
                    f"- Нормализовано дат в столбце A: {stats['dates_normalized']}\n"
                    f"- Нормализовано дат рождения в столбце C: {stats['birth_dates_normalized']}\n"
                    f"- Нормализовано дат окончания в столбце F: {stats['end_dates_normalized']}\n"
                    f"- Перемещено текстовых блоков: {stats['text_moved']}\n"
                    f"- Перемещено записей о судах: {stats['court_info_moved']}\n"
                    f"- Нормализовано дат в информации о судах: {stats['court_dates_normalized']}\n"
                    f"- Отформатировано ячеек с информацией о судах: {stats['formatted_cells']}\n"
                    f"- Всего нормализовано дат: {stats['total_dates_normalized']}\n"
                    f"- Удалены столбцы: A и C\n\n"
                    f"Обработка столбца B:\n"
                    f"- Обработано строк: {b_stats['processed_rows']}\n"
                    f"- Найдено адресов: {b_stats['addresses_found']}\n"
                    f"- Найдено телефонов: {b_stats['phones_found']}\n"
                    f"- Найдено дополнительной информации: {b_stats['other_info_found']}\n\n"
                    f"Результат сохранен в: {self.excel_path}\n\n"
                    f"Адреса перемещены в столбец O\n"
                    f"Телефоны перемещены в столбец P\n"
                    f"Прочая информация перемещена в столбец Q"
                )
                
                # Открываем Excel-файл
                self.open_file(self.excel_path)
                
                messagebox.showinfo("Успех", f"Обработка завершена. Таблицы сохранены в {self.excel_path}")
            else:
                self.update_status("В документе не найдено таблиц.")
                messagebox.showwarning("Предупреждение", "В документе не найдено таблиц")
        
        except Exception as e:
            self.update_status(f"Произошла ошибка при обработке файла:\n{str(e)}")
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
    
    def open_file(self, file_path):
        """Открытие файла в соответствующем приложении"""
        try:
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', file_path])
            else:  # Linux и другие
                subprocess.call(['xdg-open', file_path])
            return True
        except Exception as e:
            print(f"Ошибка при открытии файла: {str(e)}")
            return False

if __name__ == "__main__":
    root = tk.Tk()
    app = DocxToExcelApp(root)
    root.mainloop()
